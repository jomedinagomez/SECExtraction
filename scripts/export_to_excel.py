"""Export Content Understanding analyzer results to multi-sheet Excel workbooks.

Reads CU analyzer result JSON files (one per source PDF) from --input-dir and
writes one .xlsx per document into --output-dir. Each financial table becomes a
sheet with:

- Row 1: merged title (tableTitle + companyName + "(in <unit>)"), bold, 14pt.
- Row 2: spanning group headers (periodGroup*), merged across consecutive
  identical groups, italic. Hidden when every group is empty.
- Row 3: leaf column headers (periodHeader*), bold, bottom-bordered.
- Rows 4+: lineItem in column A with " "*3*level indentation; section headers
  bold, subtotals bold + top border.

Auto-detects nested schema (periodHeaders[]/values[]) vs portal-flat schema
(periodHeader1..6/value1..6/periodGroup1..6).

Usage:
    python scripts/export_to_excel.py --input-dir output --output-dir output/excel
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

MAX_FLAT_COLS = 6
SHEET_NAME_MAX = 31
FORBIDDEN_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
# openpyxl rejects C0 control chars (except \t, \n, \r) in cell values.
ILLEGAL_CELL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _clean(s: Any) -> Any:
    """Strip control characters that openpyxl will reject."""
    if isinstance(s, str):
        return ILLEGAL_CELL_CHARS.sub("", s)
    return s


# --------------------------------------------------------------------------- #
# CU envelope unwrapping
# --------------------------------------------------------------------------- #

def _scalar(field: dict[str, Any] | None, default: Any = "") -> Any:
    """Unwrap a CU field to its scalar value, returning `default` if missing."""
    if not field:
        return default
    for key in ("valueString", "valueInteger", "valueNumber", "valueBoolean"):
        if key in field:
            return _clean(field[key])
    return default


def _array(field: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not field:
        return []
    return field.get("valueArray", []) or []


def _obj(field: dict[str, Any] | None) -> dict[str, Any]:
    if not field:
        return {}
    return field.get("valueObject", {}) or {}


# --------------------------------------------------------------------------- #
# Schema-agnostic normalization
# --------------------------------------------------------------------------- #

def _normalize_table(raw_tbl: dict[str, Any]) -> dict[str, Any]:
    """Normalize one `financialTables[]` entry into a plain dict.

    Handles both schemas:
      * nested:  periodHeaders: array<string>, rows[].values: array<string>
      * flat:    periodHeader1..6, value1..6, periodGroup1..6
    """
    tbl_obj = _obj(raw_tbl)

    title = str(_scalar(tbl_obj.get("tableTitle"))).strip()
    company = str(_scalar(tbl_obj.get("companyName"))).strip()
    stype = str(_scalar(tbl_obj.get("statementType"))).strip() or "Other"
    unit = str(_scalar(tbl_obj.get("unit"))).strip()

    # Column headers: flat wins if any periodHeader1 exists
    period_headers: list[str] = []
    period_groups: list[str] = []
    raw_group_count = 0
    if "periodHeader1" in tbl_obj:
        for i in range(1, MAX_FLAT_COLS + 1):
            h = str(_scalar(tbl_obj.get(f"periodHeader{i}"))).strip()
            g = str(_scalar(tbl_obj.get(f"periodGroup{i}"))).strip()
            period_headers.append(h)
            period_groups.append(g)
        # Trim trailing empty columns
        while period_headers and not period_headers[-1] and not period_groups[-1]:
            period_headers.pop()
            period_groups.pop()
        raw_group_count = len(period_groups)  # flat schema: already 1:1
    else:
        period_headers = [
            str(_scalar(h)).strip() for h in _array(tbl_obj.get("periodHeaders"))
        ]
        period_groups = [
            str(_scalar(g)).strip()
            for g in _array(tbl_obj.get("periodGroupHeaders"))
        ]
        # Track original group count before padding (for header reorder fix)
        raw_group_count = len(period_groups)
        # Pad groups to match headers length
        if len(period_groups) < len(period_headers):
            period_groups += [""] * (len(period_headers) - len(period_groups))

    # Rows
    rows_norm: list[dict[str, Any]] = []
    for raw_row in _array(tbl_obj.get("rows")):
        row_obj = _obj(raw_row)
        line_item = str(_scalar(row_obj.get("lineItem"))).strip()
        if not line_item:
            continue
        level_val = _scalar(row_obj.get("level"), 0)
        try:
            level = int(level_val) if level_val != "" else 0
        except (TypeError, ValueError):
            level = 0

        is_header = bool(_scalar(row_obj.get("isSectionHeader"), False))
        is_subtotal = bool(_scalar(row_obj.get("isSubtotal"), False))
        parent = str(_scalar(row_obj.get("parentLineItem"))).strip()

        if "value1" in row_obj:
            values = [
                str(_scalar(row_obj.get(f"value{i}"))).strip()
                for i in range(1, len(period_headers) + 1)
            ]
        else:
            values = [str(_scalar(v)).strip() for v in _array(row_obj.get("values"))]

        rows_norm.append(
            {
                "lineItem": line_item,
                "level": max(level, 0),
                "parentLineItem": parent,
                "isSectionHeader": is_header,
                "isSubtotal": is_subtotal,
                "values": values,
            }
        )

    result = {
        "tableTitle": title,
        "companyName": company,
        "statementType": stype,
        "unit": unit,
        "periodGroups": period_groups,
        "periodHeaders": period_headers,
        "rows": rows_norm,
    }

    # Post-process: fix known CU ordering issues
    _fix_header_order(result, raw_group_count)
    _fix_row_order(result)

    return result


def _fix_header_order(table: dict[str, Any], raw_group_count: int) -> None:
    """Fix CU bug where grouped sub-columns are pushed to the end of periodHeaders.

    CU sometimes returns periodGroupHeaders with fewer entries than periodHeaders,
    mapping groups to the LAST N headers instead of the first. When that happens
    the values are already in correct physical left-to-right order, so we rotate
    headers (and corresponding values) to put grouped columns first.

    Detection: raw CU periodGroupHeaders count < periodHeaders count AND
    all raw group entries are non-empty (representing actual spanning headers).
    """
    hdrs = table["periodHeaders"]
    n_hdrs = len(hdrs)

    # Only applies when CU returned fewer groups than headers
    if raw_group_count == 0 or raw_group_count >= n_hdrs:
        return
    # All raw group entries must be non-empty
    grps = table["periodGroups"]
    if not all(g.strip() for g in grps[:raw_group_count]):
        return

    # The grouped columns are at the END of hdrs but should be at the START.
    # Values are already in correct physical left-to-right order (grouped cols first),
    # so we only rotate the headers to match, NOT the values.
    grouped_hdrs = hdrs[-raw_group_count:]
    ungrouped_hdrs = hdrs[:-raw_group_count]
    new_hdrs = grouped_hdrs + ungrouped_hdrs
    new_grps = grps[:raw_group_count] + [""] * len(ungrouped_hdrs)

    table["periodHeaders"] = new_hdrs
    table["periodGroups"] = new_grps


def _fix_row_order(table: dict[str, Any]) -> None:
    """Fix CU bug where child rows are displaced from their parent section header.

    CU sometimes returns all L1 rows first, then all L2 rows (sorted by level
    instead of document order). We reorder so that each section header is
    immediately followed by its children (rows at a deeper level) before the
    next sibling at the same or shallower level.

    Algorithm: scan rows for section headers. For each header at level L, collect
    all children (level > L) that reference it (by parentLineItem) or that are
    currently displaced (appear after other L-level siblings). Reinsert them
    right after the header.
    """
    rows = table["rows"]
    if len(rows) < 3:
        return

    # Build a map: header lineItem -> list of child row indices (by parentLineItem)
    header_children: dict[str, list[int]] = {}
    header_indices: dict[str, int] = {}
    for i, row in enumerate(rows):
        if row["isSectionHeader"]:
            header_children[row["lineItem"]] = []
            header_indices[row["lineItem"]] = i
        parent = row.get("parentLineItem", "")
        if parent and parent in header_children:
            header_children[parent].append(i)

    # Check if any header's children are non-contiguous (displaced)
    needs_fix = False
    for hdr_name, child_idxs in header_children.items():
        if not child_idxs:
            continue
        hdr_idx = header_indices[hdr_name]
        # Children should be at positions hdr_idx+1, hdr_idx+2, ...
        expected_start = hdr_idx + 1
        for offset, ci in enumerate(child_idxs):
            if ci != expected_start + offset:
                needs_fix = True
                break
        if needs_fix:
            break

    if not needs_fix:
        return

    # Rebuild row list: for each header, pull its children right after it
    placed = set()
    new_rows = []

    def _place_row(idx: int) -> None:
        if idx in placed:
            return
        placed.add(idx)
        row = rows[idx]
        new_rows.append(row)
        # If this is a header, place its children immediately after
        if row["isSectionHeader"] and row["lineItem"] in header_children:
            for ci in header_children[row["lineItem"]]:
                _place_row(ci)

    for i in range(len(rows)):
        _place_row(i)

    table["rows"] = new_rows


def load_document(json_path: Path) -> list[dict[str, Any]]:
    """Load a CU result file and return a list of normalized tables."""
    data = json.loads(json_path.read_text(encoding="utf-8"))
    contents = data.get("contents") or []
    if not contents:
        return []
    fields = contents[0].get("fields") or {}
    tables_field = fields.get("financialTables")
    if not tables_field:
        return []
    return [_normalize_table(t) for t in _array(tables_field)]


# --------------------------------------------------------------------------- #
# Sheet naming
# --------------------------------------------------------------------------- #

def _slug(text: str, limit: int) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_").lower()
    return s[:limit] or "table"


def sheet_name_for_table(idx: int, table: dict[str, Any], used: set[str]) -> str:
    stype = table["statementType"] or "Other"
    prefix = f"{idx:02d}_{stype}_"
    slug_budget = SHEET_NAME_MAX - len(prefix)
    base = prefix + _slug(table["tableTitle"] or stype, slug_budget)
    base = FORBIDDEN_SHEET_CHARS.sub("", base)[:SHEET_NAME_MAX]
    name = base
    n = 2
    while name in used:
        suffix = f"_{n}"
        name = base[: SHEET_NAME_MAX - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


# --------------------------------------------------------------------------- #
# Sheet writing
# --------------------------------------------------------------------------- #

_THIN = Side(style="thin", color="000000")
_BORDER_BOTTOM = Border(bottom=_THIN)
_BORDER_TOP = Border(top=_THIN)

_FONT_TITLE = Font(bold=True, size=14)
_FONT_GROUP = Font(italic=True, size=10)
_FONT_HEADER = Font(bold=True, size=11)
_FONT_SECTION = Font(bold=True)
_FONT_SUBTOTAL = Font(bold=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def _merge_consecutive(values: list[str]) -> list[tuple[int, int, str]]:
    """Return [(start_idx, end_idx, value), ...] merging consecutive equal non-empty runs."""
    spans: list[tuple[int, int, str]] = []
    i = 0
    while i < len(values):
        v = values[i]
        j = i
        while j + 1 < len(values) and values[j + 1] == v:
            j += 1
        spans.append((i, j, v))
        i = j + 1
    return spans


def write_table_sheet(ws: Worksheet, table: dict[str, Any]) -> None:
    headers = table["periodHeaders"]
    groups = table["periodGroups"]
    n_cols = len(headers)
    total_cols = 1 + n_cols  # col A = lineItem

    # --- Row 1: title ---
    bits = [b for b in (table["tableTitle"], table["companyName"]) if b]
    title_text = " — ".join(bits) if bits else (table["statementType"] or "Table")
    if table["unit"]:
        title_text += f"  (in {table['unit']})"
    ws.cell(row=1, column=1, value=title_text).font = _FONT_TITLE
    ws.cell(row=1, column=1).alignment = _ALIGN_LEFT
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.row_dimensions[1].height = 22

    # --- Row 2: group headers (only if any non-empty) ---
    data_start_row = 3
    has_groups = any(g for g in groups)
    if has_groups:
        spans = _merge_consecutive(groups)
        for start, end, grp in spans:
            if not grp:
                continue
            c = ws.cell(row=2, column=2 + start, value=grp)
            c.font = _FONT_GROUP
            c.alignment = _ALIGN_CENTER
            if end > start:
                ws.merge_cells(
                    start_row=2, start_column=2 + start,
                    end_row=2, end_column=2 + end,
                )
        data_start_row = 4

    # --- Header row: leaf periodHeaders ---
    header_row = data_start_row - 1
    c = ws.cell(row=header_row, column=1, value="Line Item")
    c.font = _FONT_HEADER
    c.alignment = _ALIGN_LEFT
    c.border = _BORDER_BOTTOM
    for i, h in enumerate(headers):
        c = ws.cell(row=header_row, column=2 + i, value=h)
        c.font = _FONT_HEADER
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_BOTTOM

    # --- Data rows ---
    r = data_start_row
    for row in table["rows"]:
        indent = "   " * row["level"]
        label_cell = ws.cell(row=r, column=1, value=f"{indent}{row['lineItem']}")
        label_cell.alignment = _ALIGN_LEFT
        if row["isSectionHeader"]:
            label_cell.font = _FONT_SECTION
        elif row["isSubtotal"]:
            label_cell.font = _FONT_SUBTOTAL
            label_cell.border = _BORDER_TOP

        vals = row["values"]
        for i in range(n_cols):
            v = vals[i] if i < len(vals) else ""
            cell = ws.cell(row=r, column=2 + i, value=v)
            cell.alignment = _ALIGN_RIGHT
            if row["isSubtotal"]:
                cell.font = _FONT_SUBTOTAL
                cell.border = _BORDER_TOP
            elif row["isSectionHeader"]:
                cell.font = _FONT_SECTION
        r += 1

    # --- Layout ---
    ws.freeze_panes = ws.cell(row=data_start_row, column=2)
    ws.column_dimensions["A"].width = 60
    for i in range(n_cols):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #

def export_document(json_path: Path, out_dir: Path) -> Path:
    tables = load_document(json_path)
    out_path = out_dir / f"{json_path.stem}.xlsx"
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove the auto-created blank sheet
    default_ws = wb.active
    wb.remove(default_ws)

    if not tables:
        ws = wb.create_sheet(title="empty")
        ws.cell(row=1, column=1, value=f"No financialTables found in {json_path.name}")
    else:
        used: set[str] = set()
        for i, tbl in enumerate(tables, start=1):
            name = sheet_name_for_table(i, tbl, used)
            ws = wb.create_sheet(title=name)
            write_table_sheet(ws, tbl)

    wb.save(out_path)
    return out_path


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input-dir", type=Path, default=Path("output"))
    ap.add_argument("--output-dir", type=Path, default=Path("output/excel"))
    ap.add_argument("--pattern", default="*.json")
    args = ap.parse_args()

    json_files = sorted(args.input_dir.glob(args.pattern))
    if not json_files:
        print(f"No JSON files found in {args.input_dir} matching {args.pattern}")
        return 1

    for jp in json_files:
        try:
            out_path = export_document(jp, args.output_dir)
            tables = load_document(jp)
            print(f"[OK] {jp.name} -> {out_path}  ({len(tables)} tables)")
        except Exception as exc:  # pragma: no cover
            print(f"[FAIL] {jp.name}: {exc}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
